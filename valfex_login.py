import tkinter
from tkinter import ttk
from ldap3 import Server, Connection, MODIFY_REPLACE, SUBTREE
import re
import getpass
import random
import requests
from ldap3.extend.microsoft.addMembersToGroups import ad_add_members_to_groups
from tkinter.messagebox import showerror, showwarning, showinfo
from ttkwidgets import CheckboxTreeview
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import webbrowser


class Application(ttk.Frame):

    def __init__(self, master=None):
        ttk.Frame.__init__(self, master)
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.s = ttk.Style()
        self.s.theme_use('clam')

        self.current_date = datetime.datetime.now()

        organization_list = ['.ru', '.ru', '.ru']
        city_list = ['Владимир', 'Вязники', 'Москва',
                     'Нижний Новгород', 'Пятигорск', 'Чебоксары']
        self.font = ("Arial Bold", 12)

        connect_ad = connect_ldap()
        connect_ad.search(search_base='OU=,DC=,DC=local', search_filter='(cn=*)',
                          search_scope=SUBTREE, attributes=['department', 'telephoneNumber', 'title', 'mobile'])

        self.department_list = []
        self.position_list = []

        for entry in connect_ad.entries:
            self.department_list.append(str(re.findall(
                r'department: ([^\r]+)', str(entry))).replace('[', '').replace(']', '').replace("'", ''))
            self.position_list.append(str(re.findall(
                r'title: ([^\r]+)', str(entry))).replace('[', '').replace(']', '').replace("'", ''))

        self.position_list = list(set(self.position_list))
        self.department_list = list(set(self.department_list))
        self.position_list.sort()
        self.department_list.sort()

        self.full_name_label = ttk.Label(
            self, text='Фамилия Имя Отчество', font=self.font)
        self.full_name_input = ttk.Entry(self, width=30, font=self.font)
        self.full_name_label.grid(row=0, column=0, sticky=tkinter.S)
        self.full_name_input.grid(row=1, column=0, sticky=tkinter.S, padx=20)

        self.position_label = ttk.Label(self, text='Должность', font=self.font)
        self.position_combobox = ttk.Combobox(
            self, width=28, font=self.font, values=self.position_list)
        self.position_combobox.bind('<KeyRelease>', self.position_check_input)
        self.position_label.grid(row=2, column=0, sticky=tkinter.S)
        self.position_combobox.grid(row=3, column=0, sticky=tkinter.S)

        self.department_label = ttk.Label(
            self, text='Подразделение', font=self.font)
        self.department_combobox = ttk.Combobox(
            self, width=28, font=self.font, values=self.department_list)
        self.department_combobox.bind(
            '<KeyRelease>', self.department_check_input)
        self.department_label.grid(row=4, column=0, sticky=tkinter.S)
        self.department_combobox.grid(row=5, column=0, sticky=tkinter.S)

        self.city_label = ttk.Label(self, text='Город', font=self.font)
        self.city_combobox = ttk.Combobox(
            self, width=20, font=self.font, values=city_list, state='readonly')
        self.city_label.grid(row=6, column=0, sticky=tkinter.S)
        self.city_combobox.grid(row=7, column=0, sticky=tkinter.S)

        self.head_label = ttk.Label(
            self, text='ФИО руководителя', font=self.font)
        self.head_input = ttk.Entry(self, width=30, font=self.font)
        self.head_label.grid(row=8, column=0, sticky=tkinter.S)
        self.head_input.grid(row=9, column=0, sticky=tkinter.S)

        self.supervisor_label = ttk.Label(
            self, text='ФИО супервайзера', font=self.font)
        self.supervisor_input = ttk.Entry(self, width=30, font=self.font)
        self.supervisor_label.grid(row=10, column=0, sticky=tkinter.S)
        self.supervisor_input.grid(row=11, column=0, sticky=tkinter.S)

        self.date_birth_label = ttk.Label(
            self, text='Дата рождения', font=self.font)
        self.date_birth_input = ttk.Entry(self, width=23, font=self.font)
        self.date_birth_label.grid(row=12, column=0, sticky=tkinter.S)
        self.date_birth_input.grid(row=13, column=0, sticky=tkinter.S)

        self.mobile_phone_label = ttk.Label(
            self, text='Мобильный телефон', font=self.font)
        self.mobile_phone_input = ttk.Entry(self, width=23, font=self.font)
        self.mobile_phone_label.grid(row=14, column=0, sticky=tkinter.S)
        self.mobile_phone_input.grid(row=15, column=0, sticky=tkinter.S)

        self.label_ad = ttk.Label(
            self, text='Расположение учетной записи', font=self.font)
        self.tree_ad = ttk.Treeview(self, height=17, selectmode='browse')
        self.vsb = ttk.Scrollbar(
            self, orient='vertical', command=self.tree_ad.yview)
        self.tree_ad.configure(yscrollcommand=self.vsb.set)
        self.tree_ad.heading('#0', text='Active Directory',
                             anchor=tkinter.W, command=self.tree_collapse_expand)
        self.tree_ad.column(column='#0', width=385)

        self.label_ad.grid(row=0, column=1, sticky=tkinter.S)
        self.tree_ad.grid(row=1, column=1, rowspan=18)
        self.vsb.grid(row=1, column=2, rowspan=18, sticky=tkinter.NSEW)

        columns = ('Фамилия Имя Отчество', 'Подразделение', 'Должность')
        self.tree_label = ttk.Label(
            self, text='Учетная запись для копирования групп', font=self.font)
        self.tree = ttk.Treeview(
            self, columns=columns, show='headings', height=5)
        self.vsb1 = ttk.Scrollbar(
            self, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.vsb1.set)
        self.tree.column(column='#1', width=225, anchor=tkinter.W,)
        self.tree.heading('Фамилия Имя Отчество', text='Фамилия Имя Отчество')
        self.tree.column(column='#2', width=265, anchor=tkinter.W, )
        self.tree.heading('Подразделение', text='Подразделение')
        self.tree.column(column='#3', width=225, anchor=tkinter.W)
        self.tree.heading('Должность', text='Должность')
        self.tree_label.grid(row=19, column=0, columnspan=2, sticky=tkinter.S)
        self.tree.grid(row=20, column=0, columnspan=3)
        self.vsb1.grid(row=20, column=3, columnspan=3, sticky=tkinter.NS)

        self.tree_ad.bind("<ButtonRelease-1>",
                          lambda *event: self.tree_update(event))
        self.tree_ad.bind("<Double-1>", lambda *event: self.tree_update(event))

        self.organization_label = ttk.Label(
            self, text='Почтовый домен', font=self.font)
        self.organization_combobox = ttk.Combobox(
            self, width=20, font=self.font, values=organization_list, state='readonly')
        self.organization_combobox.bind(
            '<<ComboboxSelected>>', self.department_label_mail_update)
        self.organization_label.grid(row=25, column=0, sticky=tkinter.S)
        self.organization_combobox.grid(row=26, column=0, sticky=tkinter.S)

        self.department_label_mail = ttk.Label(
            self, text='Подразделение почты', font=self.font)
        self.department_combobox_mail = ttk.Combobox(
            self, width=20, font=self.font, state='readonly')
        self.department_label_mail.grid(row=27, column=0, sticky=tkinter.S)
        self.department_combobox_mail.grid(row=28, column=0, sticky=tkinter.S)

        self.group_label_mail = ttk.Label(
            self, text='Группа рассылки', font=self.font)
        self.group_label_mail.grid(
            row=25, column=1, sticky=tkinter.S)

        self.group_tree = CheckboxTreeview(
            self, height=4, show='tree')
        self.group_tree.grid(
            row=26, column=1, sticky=tkinter.S, rowspan=3)
        self.group_tree.column(column='#0', width=250)

        self.mail_1c_enable = tkinter.IntVar()
        self.mail_1c_checkbutton = tkinter.Checkbutton(
            self, text='Отправка письма в Службу сопровождения 1С', font=self.font, variable=self.mail_1c_enable)
        self.mail_1c_label_company = ttk.Label(
            self, text='Организация', font=self.font)
        self.mail_1c_entry_company = ttk.Entry(self, width=30, font=self.font)
        self.mail_1c_label_bd = ttk.Label(
            self, text='Какие Базы 1С подключить', font=self.font)
        self.mail_1c_entry_bd = ttk.Entry(self, width=30, font=self.font)
        self.mail_1c_checkbutton.grid(
            row=30, column=0, columnspan=2, sticky=tkinter.S)
        self.mail_1c_label_company.grid(row=31, column=0, sticky=tkinter.S)
        self.mail_1c_entry_company.grid(row=32, column=0, sticky=tkinter.S)
        self.mail_1c_label_bd.grid(row=31, column=1, sticky=tkinter.S)
        self.mail_1c_entry_bd.grid(row=32, column=1, sticky=tkinter.S)

        self.ok_btn = tkinter.Button(self, text='Сделать хорошо', width=14, height=1, font=(
            "Arial Bold", 16), foreground='indigo', command=self.do_well)
        self.ok_btn.grid(row=40, column=0, columnspan=2,
                         sticky=tkinter.S, pady=30)

        self.update_tree_ad()
        self.full_name_input.focus_set()

    def send_mail_1c(self, full_name, login, department, position, head, supervisor, mobile, date_birth, company, bd):
        addr_from = 'it@.ru'
        addr_to = '1c@.ru'
        password = ''

        msg = MIMEMultipart()  # Создаем сообщение
        msg['From'] = addr_from  # Адресат
        msg['To'] = addr_to  # Получатель
        # Тема сообщения
        msg['Subject'] = f'Создание учетной записи в 1С для нового сотрудника {full_name}'

        body = f'''{full_name}, {position}, {department} - новый сотрудник.
{login}
Организация - {company}
Отдел сотрудника - {department}
Должность - {position}
ФИО руководителя отдела - {head}
ФИО непосредственного руководителя - {supervisor}
Какие Базы 1С подключить: {bd}
Дата рождения - {date_birth}
Мобильный телефон - {mobile}'''

        msg.attach(MIMEText(body, 'plain'))  # Добавляем в сообщение текст

        server = smtplib.SMTP_SSL('smtp.yandex.ru', 465)  # Создаем объект SMTP
        # server.starttls()             # Начинаем шифрованный обмен по TLS
        server.login(addr_from, password)  # Получаем доступ
        server.send_message(msg)  # Отправляем сообщение
        server.quit()  # Выходим

    def department_label_mail_update(self, event):
        self.department_combobox_mail.set('')
        self.department_combobox_mail.config(values='')
        for i in self.group_tree.get_children():
            self.group_tree.delete(i)

        organization = self.organization_combobox.get()
        headers = connect_yandex()
        try:
            orgId = self.search_organization_id(headers, organization)
            department_list_mail = []
            self.group_list_mail_input = []
            data = requests.get(
                url=f'https://api360.yandex.net/directory/v1/org/{orgId}/departments', headers=headers).json()
            for department_val in data['departments']:
                department_list_mail.append(department_val['name'])
            self.department_combobox_mail.config(
                values=department_list_mail)

            data = requests.get(
                url=f'https://api360.yandex.net/directory/v1/org/{orgId}/groups', headers=headers).json()
            for group_val in data['groups']:
                if group_val['type'] == 'generic':
                    self.group_list_mail_input.append(group_val['name'])

            for n, x in enumerate(self.group_list_mail_input):
                self.group_tree.insert(parent='', index='end', iid=n,
                                       text=x)
        except UnboundLocalError:
            self.organization_combobox.set('')
            msg = f'Нету прав на это домен\n\n{organization}!'
            self.show_warn(msg)
            return

    def tree_collapse_expand(self):
        for child in self.tree_ad.get_children():
            for child_0 in self.tree_ad.get_children(child):
                self.tree_ad.item(child_0, open=False)
                for child_1 in self.tree_ad.get_children(child_0):
                    self.tree_ad.item(child_1, open=False)
                    for child_2 in self.tree_ad.get_children(child_1):
                        self.tree_ad.item(child_2, open=False)
        for i in self.tree.get_children():
            self.tree.delete(i)
        for i in self.tree_ad.selection():
            self.tree_ad.selection_remove(i)

    def on_select_tree(self):

        for selected_item in self.tree.selection():
            item = self.tree.item(selected_item)
            distinguishedName = item["values"][0]

        ad_uo = self.on_select_tree_ad()
        connect_ad = connect_ldap()
        connect_ad.search(
            search_base=ad_uo, search_filter=f'(cn={distinguishedName})', search_scope='LEVEL', attributes=['memberOf'])

        list_output = []
        for entry in connect_ad.entries:
            str_entry = str(entry).replace('memberOf:', '').strip()
            lines = str_entry.split('\n')[1:]
            for line in lines:
                list_output.append(line.strip())

        return list_output

    def tree_update(self, event):

        if not self.tree_ad.selection():
            for i in self.tree.get_children():
                self.tree.delete(i)
            return

        ad_uo = self.on_select_tree_ad()

        connect_ad = connect_ldap()
        connect_ad.search(search_base=ad_uo, search_filter='(objectClass=user)',
                          search_scope='LEVEL', attributes=['department', 'title'])
        cn_list = []
        department_list = []
        title_list = []

        for i in self.tree.get_children():
            self.tree.delete(i)

        for entry in connect_ad.entries:
            cn_list.append(str(re.findall(
                r'CN=([^,]+)', str(entry))).replace('[', '').replace(']', '').replace("'", ''))
            department_list.append(str(re.findall(
                r'department: ([^\r]+)', str(entry))).replace('[', '').replace(']', '').replace("'", ''))
            title_list.append(str(re.findall(
                r'title: ([^\r]+)', str(entry))).replace('[', '').replace(']', '').replace("'", ''))

        users_list = list(zip(cn_list, department_list, title_list))

        for cn, dep, pos in users_list:
            self.tree.insert('', 'end', values=(cn, dep, pos))

    def select_group_tree(self):
        groups = []
        for item in self.group_tree.get_children():
            if 'checked' in self.group_tree.item(item, 'tags'):
                groups.append(self.group_tree.item(item, 'text'))

    def update_tree_ad(self):
        connect_ad = connect_ldap()
        # Выполняем поиск
        connect_ad.search('OU=,DC=,DC=local', '(objectClass=organizationalUnit)', attributes=[
                          'distinguishedName'], search_scope=SUBTREE)
        ou_list = []
        for entry in connect_ad.entries:
            ou_list.append(entry.distinguishedName.value)
        data = []
        # Выводим список OU
        for ou in ou_list:
            ou = ou.replace("OU=", "").replace(",DC=,DC=local", "")
            data.append(ou)

        for item in data:
            if len(item.split(',')) == 1:
                self.tree_ad.insert(
                    parent='', index=tkinter.END, text=item, open=True)

        for child in self.tree_ad.get_children():
            for item in data:
                if len(item.split(',')) == 2:
                    item1, item2 = item.split(',')
                    self.tree_ad.insert(
                        parent=child, index=tkinter.END, text=item1)
            for child_0 in self.tree_ad.get_children(child):
                for item in data:
                    if len(item.split(',')) == 3:
                        item1, item2, item3 = item.split(',')
                        if self.tree_ad.item(child_0)['text'] == item2:
                            self.tree_ad.insert(
                                parent=child_0, index=tkinter.END, text=item1)
                for child_1 in self.tree_ad.get_children(child_0):
                    for item in data:
                        if len(item.split(',')) == 4:
                            item1, item2, item3, items4 = item.split(',')
                            if self.tree_ad.item(child_1)['text'] == item2:
                                self.tree_ad.insert(
                                    parent=child_1, index=tkinter.END, text=item1)

    def on_select_tree_ad(self):
        if not self.tree_ad.selection():
            return

        item_iid = self.tree_ad.selection()
        node = self.tree_ad.item(item_iid)['text']

        parent_iid_0 = self.tree_ad.parent(item_iid)
        node_0 = self.tree_ad.item(parent_iid_0)['text']

        parent_iid_1 = self.tree_ad.parent(parent_iid_0)
        node_1 = self.tree_ad.item(parent_iid_1)['text']

        parent_iid_2 = self.tree_ad.parent(parent_iid_1)
        node_2 = self.tree_ad.item(parent_iid_2)['text']

        if node_2 != '':
            dn_ou = f'OU={node},OU={node_0},OU={node_1},OU={node_2},DC=,DC=local'
            return dn_ou
        elif node_1 != '':
            dn_ou = f'OU={node},OU={node_0},OU={node_1},DC=,DC=local'
            return dn_ou
        elif node_0 != '':
            dn_ou = f'OU={node},OU={node_0},DC=,DC=local'
            return dn_ou
        else:
            dn_ou = f'OU={node},DC=,DC=local'
            return dn_ou

    def position_check_input(self, event):
        value = event.widget.get()

        if value == '':
            self.position_combobox['values'] = self.position_list
        else:
            data = []
            for item in self.position_list:
                if value.lower() in item.lower():
                    data.append(item)

            self.position_combobox['values'] = data

    def department_check_input(self, event):
        value = event.widget.get()

        if value == '':
            self.department_combobox['values'] = self.department_list
        else:
            data = []
            for item in self.department_list:
                if value.lower() in item.lower():
                    data.append(item)

            self.department_combobox['values'] = data

    def do_well(self):

        full_name = self.full_name_input.get()
        position = self.position_combobox.get()
        department = self.department_combobox.get()
        city = self.city_combobox.get()
        head = self.head_input.get()
        supervisor = self.supervisor_input.get()
        date_birth = self.date_birth_input.get()
        mobile = self.mobile_phone_input.get()
        organization = self.organization_combobox.get()
        department_mail = self.department_combobox_mail.get()
        groups = self.select_group_tree()
        company = self.mail_1c_entry_company.get()
        bd = self.mail_1c_entry_bd.get()

        try:
            datetime.datetime.strptime(date_birth, '%d.%m.%Y')
            flg_db = 'Y'
        except Exception:
            flg_db = 'N'

        if full_name == '':
            self.full_name_input.focus_set()
            msg = 'Фамилия Имя Отчество\n\nНе заполнено!'
            self.show_warn(msg)
            return
        elif position == '':
            self.position_combobox.focus_set()
            msg = 'Должность\n\nНе заполнена!'
            self.show_warn(msg)
            return
        elif department == '':
            self.department_combobox.focus_set()
            msg = 'Подразделение\n\nНе заполнено!'
            self.show_warn(msg)
            return
        elif city == '':
            self.city_combobox.focus_set()
            msg = 'Город\n\nНе заполнен!'
            self.show_warn(msg)
            return
        elif head == '':
            self.head_input.focus_set()
            msg = 'ФИО руководителя\n\nНе заполнено!'
            self.show_warn(msg)
            return
        elif supervisor == '' and (position.lower() == 'менеджер по продажам'
                                   or position.lower() == 'помощник менеджера'
                                   or position.lower() == 'помощник менеджера по продажам'
                                   or position.lower() == 'ведущий менеджер по продажам'
                                   or position.lower() == 'ведущий менеджер'):
            self.supervisor_input.focus_set()
            msg = 'ФИО супервайзера\n\nНе заполнено!'
            self.show_warn(msg)
            return
        elif date_birth != '' and flg_db == 'N':
            self.date_birth_input.focus_set()
            msg = 'Дата некорректна\n\nФормат:\n\nDD.MM.YYYY'
            self.show_warn(msg)
            return
        elif mobile != '' and not re.match(r'\+7\d{10}', mobile):
            self.mobile_phone_input.focus_set()
            msg = 'Телефон некорректен\n\nФормат:\n\n+71234567890'
            self.show_warn(msg)
            return
        elif not self.tree_ad.selection():
            child_id = self.tree_ad.get_children()[0]
            print(child_id)
            self.tree_ad.focus(child_id)
            self.tree_ad.selection_set(child_id)
            msg = 'Путь\n\nНе выбран!'
            self.show_warn(msg)
            return
        elif not self.tree.selection():
            child_id = self.tree_ad.get_children()[0]
            print(child_id)
            self.tree_ad.focus(child_id)
            self.tree_ad.selection_set(child_id)
            msg = 'Учетная запись для групп\n\nНе выбрана!'
            self.show_warn(msg)
            return
        elif organization == '':
            self.organization_combobox.focus_set()
            msg = 'Почтовый домен\n\nНе заполнен!'
            self.show_warn(msg)
            return
        elif department_mail == '':
            self.department_combobox_mail.focus_set()
            msg = 'Подразделение почты\n\nНе заполнено!'
            self.show_warn(msg)
            return
        elif self.mail_1c_enable.get() == 1 and company == '':
            self.mail_1c_label_company.focus_set()
            msg = 'Организация\n\nНе заполнена!'
            self.show_warn(msg)
            return
        else:
            if len(full_name.split()) == 3 or len(full_name.split()) == 4:
                ad_uo = self.on_select_tree_ad()
                group = self.on_select_tree()
                last_name_rus, first_name_rus, login, middle_name_rus, password = self.login_password(
                    full_name)
                if self.create_ad(last_name_rus, first_name_rus, middle_name_rus, password, login, position, department,
                                  organization, mobile, date_birth, ad_uo, group, city) == 'Y'\
                    and self.create_mail(last_name_rus, first_name_rus, middle_name_rus, password, login,
                                         position, organization, department_mail, groups) == 'Y':
                    if self.mail_1c_enable.get() == 1:
                        self.send_mail_1c(full_name, login, department, position,
                                          head, supervisor, mobile, date_birth, company, bd)

                    data_to_excel = ([full_name, login, password, department, position, head, supervisor,
                                     organization, self.current_date, '', '', 'Активна', 'Активна', '', '', '', ''])
                    self.row_to_excel(data_to_excel)

                    self.create_widgets()
                    self.full_name_input.focus_set()

                    msg = f'Учетная запись\n{full_name}\n\nСОЗДАНА!\n\n\nДанные сохранены в\.xlsx'
                    self.show_inf(msg)
            elif len(full_name.split()) == 2:
                ad_uo = self.on_select_tree_ad()
                group = self.on_select_tree()
                last_name_rus, first_name_rus, login, password = self.login_password(
                    full_name)
                if self.create_ad_nm(last_name_rus, first_name_rus, password, login, position, department, organization,
                                     mobile, date_birth, ad_uo, group, city) == 'Y'\
                    and self.create_mail_nm(last_name_rus, first_name_rus, password, login,
                                            position, organization, department_mail, groups) == 'Y':

                    if self.mail_1c_enable.get() == 1:
                        self.send_mail_1c(full_name, login, department, position,
                                          head, supervisor, mobile, date_birth, company, bd)

                    data_to_excel = ([full_name, login, password, department, position, head, supervisor,
                                     organization, self.current_date, '', '', 'Активна', 'Активна', '', '', '', ''])
                    self.row_to_excel(data_to_excel)

                    self.create_widgets()
                    self.full_name_input.focus_set()

                    msg = f'Учетная запись\n{full_name}\n\nСОЗДАНА!\n\n\nДанные сохранены в\.xlsx'
                    self.show_inf(msg)
            else:
                msg = '\nВедите Фамилию и Имя\n\nчерез пробел'
                self.show_warn(msg)

    def row_to_excel(self, data_to_excel):
        # Путь до файла .xlsx
        path_workbook = 'Y:\\.xlsx'

        workbook = load_workbook(path_workbook)
        worksheet = workbook['main']

        for row_num in range(1, worksheet.max_row):
            if ((worksheet.cell(row=row_num, column=1).value) is None):
                cell = worksheet.cell(row=row_num, column=1)
                break

        number_cell = cell.coordinate[1:]

        char_iter = iter(('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H',
                          'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q'))
        data_iter = iter(data_to_excel)

        for val in data_iter:
            char_cell = next(char_iter)
            worksheet[f'{char_cell}{number_cell}'] = val
            worksheet[f'{char_cell}{number_cell}'].font = Font(
                name='Arial', size=10)

        workbook.save(path_workbook)
        workbook.close

    def create_mail(self, last_name_rus, first_name_rus, middle_name_rus, password, login, position, organization,
                    department_mail, groups):

        flg = 'Y'

        headers = connect_yandex()
        orgId = self.search_organization_id(headers, organization)
        departmentId = self.search_department_id(
            headers, department_mail, orgId)

        params_users = self.get_params(
            last_name_rus, first_name_rus, middle_name_rus, password, login, position, departmentId)

        # Выполнение post запроса на создание сотрудника
        response = requests.post(
            url=f'https://api360.yandex.net/directory/v1/org/{orgId}/users', headers=headers, json=params_users)

        user_full = response.json()
        userId = user_full['id']

        # Добавление сотрудника в группу рассылки
        params_members = {
            "id": userId,
            "type": 'user'
        }

        # Выполнение post запроса на добавления сотрудника в группу рассылки
        if groups is not None:
            for group in groups:
                groupId = self.search_group_id(headers, group, orgId)
                requests.post(
                    url=f'https://api360.yandex.net/directory/v1/org/{orgId}/groups/{groupId}/members', headers=headers,
                        json=params_members)

        return flg

    def create_mail_nm(self, last_name_rus, first_name_rus, password, login, position, organization,
                       department_mail, groups):

        flg = 'Y'

        headers = connect_yandex()
        orgId = self.search_organization_id(headers, organization)
        departmentId = self.search_department_id(
            headers, department_mail, orgId)
        params_users = self.get_params_nm(
            last_name_rus, first_name_rus, password, login, position, departmentId)

        # Выполнение post запроса на создание сотрудника
        response = requests.post(
            url=f'https://api360.yandex.net/directory/v1/org/{orgId}/users', headers=headers, json=params_users)

        user_full = response.json()
        userId = user_full['id']

        # Добавление сотрудника в группу рассылки
        params_members = {
            "id": userId,
            "type": 'user'
        }

        # Выполнение post запроса на добавления сотрудника в группу рассылки
        if groups is not None:
            for group in groups:
                groupId = self.search_group_id(headers, group, orgId)
                requests.post(
                    url=f'https://api360.yandex.net/directory/v1/org/{orgId}/groups/{groupId}/members', headers=headers,
                        json=params_members)

        return flg

    def get_params(self, last_name_rus, first_name_rus, middle_name_rus, password, login, position, departmentId):
        password_mail = password+'@'
        params_users = {
            "departmentId": departmentId,
            "name": {
                "first": first_name_rus,
                "last": last_name_rus,
                "middle": middle_name_rus
            },
            "nickname": login,
            "password": password_mail,
            "position": position,
        }
        return params_users

    def get_params_nm(self, last_name_rus, first_name_rus, password, login, position, departmentId):
        password_mail = password+'@'
        params_users = {
            "departmentId": departmentId,
            "name": {
                "first": first_name_rus,
                "last": last_name_rus
            },
            "nickname": login,
            "password": password_mail,
            "position": position,
        }
        return params_users

    # Функция определение id организации (домен)
    def search_organization_id(self, headers, organization):
        data = requests.get(
            url='https://api360.yandex.net/directory/v1/org', headers=headers).json()
        for organization_val in data['organizations']:
            if organization in organization_val.values():
                orgId = list(organization_val.values())[0]
        return orgId

    # Функция определение id подразделения
    def search_department_id(self, headers, department_mail, orgId):
        data = requests.get(
            url=f'https://api360.yandex.net/directory/v1/org/{orgId}/departments', headers=headers).json()
        for department_val in data['departments']:
            if department_mail in department_val.values():
                departmentId = list(department_val.values())[0]
        return departmentId

    # Функция определение id группы рассылки
    def search_group_id(self, headers, group, orgId):
        data = requests.get(
            url=f'https://api360.yandex.net/directory/v1/org/{orgId}/groups', headers=headers).json()
        for groups_val in data['groups']:
            if group in groups_val.values():
                groupId = list(groups_val.values())[0]
        return groupId

    def create_ad(self, last_name_rus, first_name_rus, middle_name_rus, password, login, position, department,
                  organization, mobile, date_birth, ad_uo, group, city):

        connect_ad = connect_ldap()

        distinguishedName = f'cn={last_name_rus} {first_name_rus} {middle_name_rus},{ad_uo}'

        attributes = self.get_attributes(first_name_rus, last_name_rus, middle_name_rus,
                                         login, position, department, organization, mobile, date_birth, city)

        flg = 'Y'

        # Создание учетной записи
        if not connect_ad.add(dn=distinguishedName, object_class='user', attributes=attributes):
            flg = 'N'
            if connect_ad.result.get("description") == 'entryAlreadyExists':
                msg = f'''Ошибка создания пользователя\n{
                    last_name_rus} {first_name_rus} {middle_name_rus}:
                    \n\nЛогин {login} уже существует!'''
                self.show_error(msg)
            return flg

        # Устанавка пароля
        try:
            connect_ad.extend.microsoft.modify_password(
                user=distinguishedName, new_password=password, old_password=None)
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nОшибка создания пароля'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Разблокировка учётной записи
        try:
            connect_ad.extend.microsoft.unlock_account(user=distinguishedName)
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nНе удалось разблокировать учетную запись'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Устанавка атрибутов
        try:
            connect_ad.modify(distinguishedName, {
                              'userAccountControl': [(MODIFY_REPLACE, [66048])]})
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nНе удалось поменять атрибуты'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Добавление в группу

        for gr in group:
            try:
                ad_add_members_to_groups(connect_ad, distinguishedName, gr)
            except Exception:
                connect_ad.delete(distinguishedName)
                msg = '\n\nНе удалось добавить группу'
                self.show_error(msg)
                flg = 'N'
                return flg

        connect_ad.unbind()
        return flg

    def create_ad_nm(self, last_name_rus, first_name_rus, password, login, position, department, organization, mobile,
                     date_birth, ad_uo, group, city):

        connect_ad = connect_ldap()

        distinguishedName = f'cn={last_name_rus} {first_name_rus},{ad_uo}'

        attributes = self.get_attributes_nm(
            first_name_rus, last_name_rus, login, position, department, organization, mobile, date_birth, city)

        flg = 'Y'

        # Создание учетной записи
        if not connect_ad.add(dn=distinguishedName, object_class='user', attributes=attributes):
            flg = 'N'
            if connect_ad.result.get("description") == 'entryAlreadyExists':
                msg = f'Ошибка создания\n{last_name_rus} {first_name_rus}:\n\nЛогин {login} уже существует!'
                self.show_error(msg)
            return flg

        # Устанавка пароля
        try:
            connect_ad.extend.microsoft.modify_password(
                user=distinguishedName, new_password=password, old_password=None)
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nОшибка создания пароля'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Разблокировка учётной записи
        try:
            connect_ad.extend.microsoft.unlock_account(user=distinguishedName)
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nНе удалось разблокировать учетную запись'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Устанавка атрибутов
        try:
            connect_ad.modify(distinguishedName, {
                              'userAccountControl': [(MODIFY_REPLACE, [66048])]})
        except Exception:
            connect_ad.delete(distinguishedName)
            msg = '\n\nНе удалось поменять атрибуты'
            self.show_error(msg)
            flg = 'N'
            return flg

        # Добавление в группу
        for gr in group:
            try:
                ad_add_members_to_groups(connect_ad, distinguishedName, gr)
            except Exception:
                connect_ad.delete(distinguishedName)
                msg = '\n\nНе удалось добавить группу'
                self.show_error(msg)
                flg = 'N'
                return flg

        connect_ad.unbind()
        return flg

    def get_attributes(self, first_name_rus, last_name_rus, middle_name_rus, login, position, department, organization,
                       mobile, date_birth, city):
        if mobile == '' and date_birth == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus} {middle_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                'initials': middle_name_rus[0],  # Инициалы
                'middleName': middle_name_rus,  # Отчество
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}'  # E-mail
            }
        elif date_birth == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus} {middle_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                'initials': middle_name_rus[0],  # Инициалы
                'middleName': middle_name_rus,  # Отчество
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}',  # E-mail
                            'mobile': f'{mobile}'  # Дата рождения
            }
        elif mobile == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus} {middle_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                'initials': middle_name_rus[0],  # Инициалы
                'middleName': middle_name_rus,  # Отчество
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}',  # E-mail
                            'info': f'{date_birth}',  # Дата рождения
            }
        else:
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus} {middle_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                'initials': middle_name_rus[0],  # Инициалы
                'middleName': middle_name_rus,  # Отчество
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город,
                            'mail': f'{login}@{organization}',  # E-mail
                            'info': f'{date_birth}',  # Дата рождения
                            'mobile': f'{mobile}',  # Мобильный телефон
                            # 'telephoneNumber': #Внутренний номер
                            # 'physicalDeliveryOfficeName': #Офис
                            # 'jpegPhoto': #Фото
                            # 'thumbnailPhoto': #Фото2
            }
        return attributes

    def get_attributes_nm(self, first_name_rus, last_name_rus, login, position, department, organization, mobile,
                          date_birth, city):
        if mobile == '' and date_birth == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}'  # E-mail
            }
        elif date_birth == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}',  # E-mail
                            'mobile': f'{mobile}'  # Дата рождения
            }
        elif mobile == '':
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}',  # E-mail
                            'info': f'{date_birth}',  # Дата рождения
            }
        else:
            attributes = {
                # Отображаемое имя
                'displayName': f'{last_name_rus} {first_name_rus}',
                'givenName': first_name_rus,  # Имя
                'sn': last_name_rus,  # Фамилия
                # Имя входа пользователя
                'userPrincipalName': f'{login}@.local',
                # Имя входа пользователя (пред-Windows 2000)
                'sAMAccountName': login,
                'title': position,  # Должность
                'department': department,  # Подразделение
                'l': city,  # Город
                            'mail': f'{login}@{organization}',  # E-mail
                            'info': f'{date_birth}',  # Дата рождения
                            'mobile': f'{mobile}',  # Мобильный телефон
                            # 'telephoneNumber': #Внутренний номер
                            # 'physicalDeliveryOfficeName': #Офис
                            # 'jpegPhoto': #Фото
                            # 'thumbnailPhoto': #Фото2
            }
        return attributes

    def login_password(self, full_name):
        # ФИО по отдельности кириллицей
        if len(full_name.split()) == 3:
            last_name_rus, first_name_rus, middle_name_rus = full_name.split()
            # Транслитерация ФИО
            transliterated_text = self.translit(full_name)
            # ФИО по отдельности латиницей
            last_name, first_name, middle_name = transliterated_text.split()
            # Формирования логина
            login = f"{last_name.lower()}_{first_name[0].lower()}{middle_name[0].lower()}"
            # Формирования пароля
            password_random = f"{last_name[0]}{first_name[0]}{middle_name[0]}"
            random_numbers = self.generate_random_number(3)
            random_letters = self.generate_random_string(3)
            password = f"{password_random.lower()}{random_numbers.lower()}{random_letters}"
            print(password)
            return last_name_rus, first_name_rus, login, middle_name_rus, password
        elif len(full_name.split()) == 2:
            last_name_rus, first_name_rus = full_name.split()
            # Транслитерация ФИО
            transliterated_text = self.translit(full_name)
            # ФИО по отдельности латиницей
            last_name, first_name = transliterated_text.split()
            # Формирования логина
            login = f'{last_name.lower()}_{first_name[0].lower()}'
            # Формирования пароля
            password_random = f'{last_name[0]}{first_name[0]}'
            random_numbers = self.generate_random_number(3)
            random_letters = self.generate_random_string(3)
            password = f"{password_random.lower()}{random_numbers.lower()}{random_letters}"
            print(password)
            return last_name_rus, first_name_rus, login, password
        elif len(full_name.split()) == 4:
            last_name_rus, first_name_rus, middle_name_rus, other_name_rus = full_name.split()
            middle_name_rus = f'{middle_name_rus} {other_name_rus}'
            transliterated_text = self.translit(full_name)
            # ФИО по отдельности латиницей
            last_name, first_name, middle_name, other_name = transliterated_text.split()
            middle_name = f'{middle_name} {other_name}'
            # Формирования логина
            login = f"{last_name.lower()}_{first_name[0].lower()}{middle_name[0].lower()}"
            # Формирования пароля
            password_random = f"{last_name[0]}{first_name[0]}{middle_name[0]}"
            random_numbers = self.generate_random_number(3)
            random_letters = self.generate_random_string(3)
            password = f"{password_random.lower()}{random_numbers.lower()}{random_letters}"
            print(password)
            return last_name_rus, first_name_rus, login, middle_name_rus, password
        else:
            msg = '\nВедите Фамилию и Имя\n\nчерез пробел'
            self.show_warn(msg)

    # Функция для траслита

    def translit(self, full_name):
        translit_dict = {
            'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i',
            'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't',
            'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ъ': '', 'ы': 'y', 'ь': '',
            'э': 'e', 'ю': 'yu', 'я': 'ya',
            'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo', 'Ж': 'Zh', 'З': 'Z', 'И': 'I',
            'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T',
            'У': 'U', 'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch', 'Ъ': '', 'Ы': 'Y', 'Ь': '',
            'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
        }
        result = ''
        for char in full_name:
            if char in translit_dict:
                result += translit_dict[char]
            else:
                result += char
        return result

    # Функция генерации случайных латинских букв
    def generate_random_string(self, length):
        letters = "abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVXYZ"
        return "".join(random.choice(letters) for _ in range(length))

    # Функция генерации случайных чисел
    def generate_random_number(self, length):
        return str(random.randint(0, 10**length-1))

    # Вывод ошибки
    def show_error(self, msg):
        showerror(title=f'{getpass.getuser()}', message=msg)

    # Вывод предуприждения
    def show_warn(self, msg):
        showwarning(title=f'{getpass.getuser()}', message=msg)

    # Вывод информации
    def show_inf(self, msg):
        showinfo(title=f'{getpass.getuser()}', message=msg)


class PassApplication(ttk.Frame):
    def __init__(self, master=None):
        ttk.Frame.__init__(self, master)
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.user_label = tkinter.Label(
            self, text=f'{getpass.getuser()}', font=("Arial Bold", 30))
        self.pass_input = tkinter.Entry(
            self, show='*', width=14, font=("Arial Bold", 20))
        self.pass_btn = tkinter.Button(self, text='Вход', command=self.pass_check, width=12, height=1, font=(
            "Arial Bold", 21), foreground='indigo')
        self.user_label.grid(row=1, column=0, columnspan=1,
                             padx=5, pady=12, sticky=tkinter.S)
        self.pass_input.grid(row=2, column=0, columnspan=2,
                             padx=10, pady=5, sticky=tkinter.S)
        self.pass_btn.grid(row=3, column=0, columnspan=2,
                           padx=10, pady=12, sticky=tkinter.S)
        self.pass_input.bind("<Return>", self.pass_check)
        self.pass_input.focus_set()

    def pass_check(self, *args):
        global pass_ad
        pass_ad = self.pass_input.get()
        try:
            connect_ldap()

            pass_root.destroy()

            webbrowser.open(
                'https://oauth.yandex.ru/authorize?response_type=code&client_id=')

            global code_root
            code_root = tkinter.Tk()
            code_root.title('Код Yandex')
            w, h = 300, 200
            code_root.geometry(
                f"{w}x{h}+{(code_root.winfo_screenwidth()-w)//2}+{(code_root.winfo_screenheight()-h)//2}")
            code_root.resizable(width=False, height=False)
            code_app = AccessApplication(master=code_root)
            code_app.mainloop()
        except Exception:
            self.pass_input.delete(0, tkinter.END)
            msg = 'Неправильный пароль'
            showerror(title=f'{getpass.getuser()}', message=msg)


class AccessApplication(ttk.Frame):
    def __init__(self, master=None):
        ttk.Frame.__init__(self, master)
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.code_label = tkinter.Label(
            self, text='Код Yandex', font=("Arial Bold", 30))
        self.code_input = tkinter.Entry(
            self, width=14, font=("Arial Bold", 20))
        self.code_btn = tkinter.Button(self, text='Вход', command=self.code_check, width=12, height=1, font=(
            "Arial Bold", 21), foreground='indigo')
        self.code_label.grid(row=1, column=0, columnspan=1,
                             padx=5, pady=12, sticky=tkinter.S)
        self.code_input.grid(row=2, column=0, columnspan=2,
                             padx=10, pady=5, sticky=tkinter.S)
        self.code_btn.grid(row=3, column=0, columnspan=2,
                           padx=10, pady=12, sticky=tkinter.S)
        self.code_input.bind("<Return>", self.code_check)
        self.code_input.focus_set()

    def code_check(self, *args):
        global code
        code = self.code_input.get()
        try:
            get_access_token()
            code_root.destroy()
            root = tkinter.Tk()
            root.title(
                f'Создание учетной записи        Пользователь: {getpass.getuser()}')
            w, h = 750, 800
            root.geometry(
                f"{w}x{h}+{(root.winfo_screenwidth()-w)//2}+{(root.winfo_screenheight()-h)//2}")
            root.resizable(width=False, height=False)
            app = Application(master=root)
            app.full_name_input.focus_force()
            app.mainloop()
        except Exception:
            self.code_input.delete(0, tkinter.END)
            msg = 'Неправильный код Yandex'
            showerror(title='Неправильный код Yandex', message=msg)


def connect_ldap():
    server_ad = Server(host='', port=636, use_ssl=True)
    return Connection(server_ad, user=f'\\{getpass.getuser()}', password=pass_ad, auto_bind=True)


def get_access_token():

    # Запрос токена доступа
    token_url = 'https://oauth.yandex.ru/token'
    data = {
        'code': code,
        'client_id': '',
        'client_secret': '',
        'grant_type': 'authorization_code'
    }
    response = requests.post(token_url, data=data)
    global access_token
    access_token = response.json()['access_token']


def connect_yandex():
    # параметры headers
    headers = {'Content-Type': 'application/json',
               'Accept': 'application/json', 'Authorization': f'OAuth {access_token}'}
    return headers


pass_root = tkinter.Tk()
pass_root.title(f'{getpass.getuser()}')
w, h = 300, 200
pass_root.geometry(
    f"{w}x{h}+{(pass_root.winfo_screenwidth()-w)//2}+{(pass_root.winfo_screenheight()-h)//2}")
pass_root.resizable(width=False, height=False)
pass_app = PassApplication(master=pass_root)
pass_app.mainloop()
